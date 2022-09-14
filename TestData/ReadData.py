import openpyxl
wk = openpyxl.load_workbook("E:\\TestData100.xlsx")

def fetch_number_of_rows(Sheetname):
    sh = wk[Sheetname]
    return sh.max_row
#print(sh.max_row)
#cell = sh.cell(1,2)
#print(cell.value)
print(fetch_number_of_rows("sheet1"))